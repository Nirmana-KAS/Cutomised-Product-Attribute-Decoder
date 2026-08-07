"""
Customised Product Spec Decoder
===============================
A tkinter desktop application that decodes customised customer product
orders (invoice-style Excel files) into ERP (Odoo) product template
format. Takes human-readable customer invoice rows (stone name, shape,
size, polish plan references) and converts them into standard product
template exports with 8-digit SSKU name codes.

ORAVA (Pvt) Ltd | Developed by Shehan Nirmana | v1.3 — 2026

Screens
    1. Library Manager       - manage the 8 JSON reference libraries
    2. Upload & Process      - load the invoice file, decode each row
    3. Export                - 26-column individual-product Odoo export

Key rules (v1.3)
    * every product is FINISHED — the Name code starts with "F" (8 chars)
    * every product is INDIVIDUAL — no grouping, no comma-separated values
    * Internal Reference comes ONLY from the Ref.No column
    * Mixed / MX / MIXED sizes are skipped (no BOM can be defined)
    * six named projects share a Ref.No — Plan No. differentiates them

All library data lives in ``app_data`` as JSON and is always loaded fresh
from disk before any processing run.
"""

import os
import re
import sys
import json
import math
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl import load_workbook, Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

try:
    from PIL import Image, ImageTk
    PIL_AVAILABLE = True
except Exception:  # pragma: no cover
    PIL_AVAILABLE = False


# --------------------------------------------------------------------------
# Constants / theme
# --------------------------------------------------------------------------
APP_TITLE = "Customised Product Spec Decoder"
APP_VERSION = "v1.3 — 2026"
FOOTER_TEXT = f"ORAVA (Pvt) Ltd | Developed by Shehan Nirmana | {APP_VERSION}"

BG = "#FFFFFF"          # window background
CARD = "#F0F2F5"        # card background
TEXT = "#1a1a1a"        # primary text
DIM = "#666666"         # dim / secondary text
HIGHLIGHT = "#e94560"   # accent / highlight
SUCCESS = "#2ecc71"     # success (green export — expanded)
SUCCESS_TEXT = "#1a7f37"  # darker green — readable as text on light backgrounds
INFO = "#2980b9"        # info (blue export — compact)
WARNING = "#e67e22"     # warning (orange export)
DANGER = "#c0392b"      # delete

FONT = "Segoe UI"
F_TITLE = (FONT, 20, "bold")
F_HEAD = (FONT, 14, "bold")
F_SUB = (FONT, 11, "bold")
F_BODY = (FONT, 10)
F_SMALL = (FONT, 8)

# Export styling
XL_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
XL_ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
XL_WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
XL_HEADER_FONT = Font(name="Aptos Narrow", bold=True, color="FFFFFF")
XL_DATA_FONT = Font(name="Aptos Narrow", color="000000")


def resource_path(rel_name):
    """Absolute path to a bundled resource (works for PyInstaller frozen exe)."""
    if hasattr(sys, "_MEIPASS"):
        candidate = os.path.join(sys._MEIPASS, rel_name)
        if os.path.exists(candidate):
            return candidate
    base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, rel_name)


# Persistent storage location for the library JSON files.
#
# For a frozen --onefile exe, ``__file__`` (and ``sys._MEIPASS``) point into a
# temporary extraction folder that PyInstaller deletes on exit — writing the
# libraries there loses them every close. Anchor app_data next to the *real*
# executable instead (or next to main.py in dev mode). ``sys._MEIPASS`` is used
# only for read-only bundled assets (logo.png / icon.ico) via resource_path().
if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

APP_DIR = BASE_DIR
DATA_DIR = os.path.join(BASE_DIR, "app_data")


# --------------------------------------------------------------------------
# Odoo export column definitions
# --------------------------------------------------------------------------
COL_REF = "Internal Reference"
COL_NAME = "Name"
COL_CATEGORY = "Product Category"
COL_TAGS = "Tags"
COL_UNIT = "Unit"
COL_ATTR = "Product Attributes / Attribute"
COL_VALUES = "Product Attributes / Values"
COL_PLAN = "Plan No."
COL_REMARKS = "Remarks"
COL_SSKU = "SSKU with Size"
COL_LAB = "LAB Certification"
COL_ORIGIN = "Origin"
COL_UPLOAD_NAME = "Upload Ref Sheet Name"
COL_UPLOAD_ROW = "Upload Ref Sheet Raw No."

# Screen 3 — the single 26-column export layout (one product per entry).
PRODUCT_HEADERS = [
    COL_REF, COL_NAME, COL_CATEGORY, COL_TAGS, COL_UNIT, COL_ATTR, COL_VALUES,
    "Sales", "Purchase", "Product Type", "Invoicing Policy",
    "Control Policy", "Inventory tracking", "Tracking",
    "Sales Price", "Sales Taxes", "Cost", "Purchase Taxes",
    "Company", COL_SSKU, COL_PLAN, COL_REMARKS, COL_LAB, COL_ORIGIN,
    COL_UPLOAD_NAME, COL_UPLOAD_ROW,
]

SKIPPED_HEADERS = [
    "Row Number", "Upload Filename", "Item No.", "Order No.", "Ref.No",
    "LAB Certification", "Origin", "Variety", "Shape", "Size", "Height",
    "Plan No.", "Remarks", "Skip Reason",
]

# Max rows rendered in the on-screen previews (the full set is always
# streamed to the Excel export, however large).
PREVIEW_LIMIT = 2000

# The six named projects. Their upload rows share a single Ref.No, so the
# Plan No. is appended to keep each product's Internal Reference unique
# (and it also joins the product key used for duplicate resolution).
PROJECT_NAMES = ["#7-Nautilus", "OVS Boucle", "OVS Lunette", "No. 3",
                 "Tache", "Gala"]
_PROJECT_LOOKUP = {p.upper(): p for p in PROJECT_NAMES}

# Size values that cannot produce a BOM and are skipped outright.
MIXED_SIZES = {"MIXED", "MX"}
INVALID_SIZES = {"FS"}

# Variety values that are not stones at all.
INVALID_VARIETIES = {"METAL"}

# Constant boilerplate values shared by every product record.
BOILER = {
    COL_UNIT: "Units",
    "Sales": "True",
    "Purchase": "True",
    "Product Type": "Goods",
    "Invoicing Policy": "Delivered quantities",
    "Control Policy": "on received quantity",
    "Inventory tracking": "True",
    "Tracking": "By  Lot",
    "Sales Price": "",
    "Sales Taxes": "",
    "Cost": "",
    "Purchase Taxes": "",
    "Company": "Orava (Pvt) Ltd",
}

# Explicit column widths (others auto-sized).
COL_WIDTHS = {
    COL_NAME: 14,
    COL_CATEGORY: 45,
    COL_UNIT: 8,
    COL_TAGS: 40,
    COL_ATTR: 25,
    COL_VALUES: 50,
    COL_LAB: 25,
    COL_ORIGIN: 20,
    COL_UPLOAD_NAME: 30,
    COL_UPLOAD_ROW: 15,
}


# --------------------------------------------------------------------------
# Library definitions (libraries 1-7 — standard CRUD cards)
# --------------------------------------------------------------------------
LIBRARIES = {
    "stone": {
        "title": "Stone Name Library",
        "desc": "1-4 digit stone codes → stone names (~115 entries).",
        "file": "stone_library.json",
        "columns": [("1-4 Digit", "code"), ("Stone name", "name")],
        "key_fields": ("code", "name"), "grouped": False},
    "shape": {
        "title": "Shape Library (2-Digit)",
        "desc": "2-digit shape codes → shape names + size type (L & W / D & H).",
        "file": "shape_library.json",
        "columns": [("Shape Code", "code"), ("Shape Name", "name"),
                    ("Size Type", "size_type")],
        "key_fields": ("code",), "grouped": False},
    "fullname_shape": {
        "title": "Full Name Shape Library",
        "desc": "Long / variant shape names → 2-digit shape codes.",
        "file": "fullname_shape_library.json",
        "columns": [("Shape", "shape"), ("Shape Code", "code"),
                    ("Shape Name", "name"), ("Size Type", "size_type")],
        "key_fields": ("shape",), "grouped": False},
    "polish": {
        "title": "Polish Type Lookup Library",
        "desc": "Shape + Ref.No + Plan No. + Remarks → polish type (~1,126 entries).",
        "file": "polish_type_library.json",
        "columns": [("Shape", "shape"), ("Ref.No", "ref_no"),
                    ("Plan No.", "plan_no"), ("Remarks", "remarks"),
                    ("Polish Type Code", "code"),
                    ("Polishing Type Name", "name")],
        "key_fields": ("shape", "ref_no", "plan_no", "remarks", "code", "name"),
        "grouped": False},
    "origin": {
        "title": "Origin Library",
        "desc": "Origin codes → origin names (~23 entries).",
        "file": "origin_library.json",
        "columns": [("Origin Code", "code"), ("Origin Name", "name")],
        "key_fields": ("code",), "grouped": False},
    "category": {
        "title": "Category Name Library",
        "desc": "Stone name → parent category paths (Finish paths, grouped).",
        "file": "category_library.json",
        "columns": [("Stone Name", "stone"), ("Parent Category", "parent")],
        "key_fields": ("stone", "parent"), "grouped": True},
    "colour": {
        "title": "Colour Library",
        "desc": "Stone name → colours (reference only — Colour stays blank).",
        "file": "colour_library.json",
        "columns": [("Stone Name", "stone"), ("Colour", "colour")],
        "key_fields": ("stone", "colour"), "grouped": True},
}

# Library 8 — Previously Decoded Products (special card, JSON list of dicts)
DECODED_FILE = "decoded_products_library.json"
DECODED_COLUMNS = [
    ("Internal Reference", "internal_ref"), ("Name", "name_code"),
    ("Product Category", "category"), ("Tags", "tags"),
    ("Stone Name", "stone"), ("Polish Type Code", "polish_code"),
    ("Polishing Type Name", "polish_name"), ("Shape Code", "shape_code"),
    ("Shape Name", "shape_name"), ("Size Type", "size_type"),
    ("Size", "size"), ("Height", "height"),
    ("SSKU with Size", "ssku"),
    ("Origin Code", "origin_code"), ("Origin Name", "origin_name"),
    ("Plan No.", "plan_no"), ("Remarks", "remarks"),
    ("LAB Certification", "lab_certification"),
    ("Origin", "origin_display"),
    ("Upload Ref Sheet Name", "upload_filename"),
    ("Upload Ref Sheet Raw No.", "upload_row_number"),
]


# --------------------------------------------------------------------------
# Data layer (JSON load / save)
# --------------------------------------------------------------------------
def ensure_data_dir():
    if not os.path.isdir(DATA_DIR):
        os.makedirs(DATA_DIR, exist_ok=True)


def lib_path(lib_key):
    return os.path.join(DATA_DIR, LIBRARIES[lib_key]["file"])


def _load_json_list(path):
    if not os.path.exists(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        return data if isinstance(data, list) else []
    except (json.JSONDecodeError, OSError):
        return []


def _save_json_list(path, rows):
    ensure_data_dir()
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(rows, fh, indent=2, ensure_ascii=False)


def load_rows(lib_key):
    """Load a library's normalised rows fresh from disk."""
    ensure_data_dir()
    return _load_json_list(lib_path(lib_key))


def save_rows(lib_key, rows):
    _save_json_list(lib_path(lib_key), rows)


def decoded_path():
    return os.path.join(DATA_DIR, DECODED_FILE)


def load_decoded():
    ensure_data_dir()
    return _load_json_list(decoded_path())


def save_decoded(rows):
    _save_json_list(decoded_path(), rows)


# --------------------------------------------------------------------------
# Value cleaning helpers
# --------------------------------------------------------------------------
def clean(value):
    """Normalise a cell value: str, \\xa0 → space, strip. None → ''.

    Floats that are whole numbers lose their trailing '.0' so Order No.
    12345.0 reads back as '12345' and size 4.0 as '4'.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).replace("\xa0", " ").strip()
    # collapse doubled internal spaces introduced by \xa0 replacement
    return text.strip()


def norm(value):
    """Case-insensitive comparison form of a cleaned value."""
    return clean(value).upper()


def is_number(text):
    try:
        float(text)
        return True
    except (TypeError, ValueError):
        return False


_RANGE_RE = re.compile(r"^(\d+(?:\.\d+)?)\s*-\s*(\d+(?:\.\d+)?)$")


def format_size_value(value):
    """Format any size number to exactly 2 decimal places.

    '4' → '4.00', '2.2' → '2.20', '1.83' → '1.83'.
    Ranges format both parts: '1.6-1.69' → '1.60-1.69'.
    'Mixed' → 'Mixed'; None / blank → ''; other text unchanged.
    """
    s = clean(value)
    if not s:
        return ""
    if s.lower() == "mixed":
        return "Mixed"
    m = _RANGE_RE.match(s)
    if m:
        return f"{float(m.group(1)):.2f}-{float(m.group(2)):.2f}"
    if is_number(s):
        return f"{float(s):.2f}"
    return s


def size_to_range(value):
    """Convert any size value to the company standard floor-based range.

    Single value → floor to integer, 'X.00-X.99' (1.83 → '1.00-1.99',
    10 → '10.00-10.99'). Range → floor of the FIRST number
    ('5.34-5.80' → '5.00-5.99'). 'Mixed', blank or non-numeric → ''.
    """
    s = clean(value)
    if not s or s.lower() == "mixed":
        return ""
    m = _RANGE_RE.match(s)
    num = m.group(1) if m else s
    if is_number(num):
        base = math.floor(float(num))
        return f"{base}.00-{base}.99"
    return ""


def uniq(values):
    """Order-preserving unique non-blank values."""
    return list(dict.fromkeys(v for v in values if v))


# --------------------------------------------------------------------------
# Excel import / export (libraries)
# --------------------------------------------------------------------------
_HEADER_WORDS = {
    "1-4 digit", "stone name", "stone", "shape code", "shape name",
    "size type", "shape", "ref.no", "plan no.", "remarks",
    "polish type code", "polishing type name", "origin code", "origin name",
    "parent category", "colour", "code", "name",
}


def import_excel(lib_key, path):
    """Read Excel into normalised rows and append ALL of them (no dedup).

    The libraries intentionally contain duplicate / similar entries with
    spelling variations, so nothing is removed. Handles the grouped
    (fill-down first column) layout for Category / Colour.
    Returns (merged_rows, added, skipped) — skipped is always 0.
    """
    cfg = LIBRARIES[lib_key]
    keys = [k for _, k in cfg["columns"]]

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    incoming = []
    last_group_value = None
    for r_idx, raw in enumerate(ws.iter_rows(values_only=True)):
        if raw is None:
            continue
        if r_idx == 0:  # skip a header row if present
            texts = {clean(c).lower() for c in raw if c is not None}
            if texts & _HEADER_WORDS:
                continue

        cells = list(raw)
        row = {}
        for i, k in enumerate(keys):
            val = cells[i] if i < len(cells) else None
            row[k] = clean(val)

        if cfg["grouped"]:
            if row[keys[0]]:
                last_group_value = row[keys[0]]
            else:
                row[keys[0]] = last_group_value or ""

        if not any(row[k] for k in keys):
            continue
        if cfg["grouped"] and not row[keys[1]]:
            continue
        incoming.append(row)
    wb.close()

    # Libraries 1-7 intentionally keep duplicate / similar entries (spelling
    # variations catch messy upload data) — ALL rows are appended, no dedup.
    merged = load_rows(lib_key) + incoming
    save_rows(lib_key, merged)
    return merged, len(incoming), 0


def export_library_excel(lib_key, path):
    cfg = LIBRARIES[lib_key]
    rows = load_rows(lib_key)
    headers = [disp for disp, _ in cfg["columns"]]
    data = [[row.get(k, "") for _, k in cfg["columns"]] for row in rows]
    write_styled_workbook(path, headers, data, sheet_name=cfg["title"][:31])


def export_decoded_excel(path):
    rows = load_decoded()
    headers = [disp for disp, _ in DECODED_COLUMNS]
    data = [[row.get(k, "") for _, k in DECODED_COLUMNS] for row in rows]
    write_styled_workbook(path, headers, data, sheet_name="Decoded Products")


# --------------------------------------------------------------------------
# Styled workbook writer (shared by all exports)
# --------------------------------------------------------------------------
def write_styled_workbook(path, headers, rows, sheet_name="Sheet1"):
    """Write a styled workbook: blue header, alternating rows, frozen header."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = XL_HEADER_FILL
        cell.font = XL_HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)

    for i, row in enumerate(rows):
        ws.append(list(row))
        fill = XL_ALT_FILL if i % 2 else XL_WHITE_FILL
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.font = XL_DATA_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    # Column widths
    for idx, header in enumerate(headers, start=1):
        letter = get_column_letter(idx)
        if header in COL_WIDTHS:
            ws.column_dimensions[letter].width = COL_WIDTHS[header]
        else:
            longest = len(str(header))
            for row in rows:
                if idx - 1 < len(row):
                    longest = max(longest, len(str(row[idx - 1])))
            ws.column_dimensions[letter].width = min(max(longest + 2, 10), 32)

    ws.freeze_panes = "A2"
    wb.save(path)


# Fixed column widths for the streamed export (write-only mode cannot rescan
# data to auto-size, so widths are chosen up front).
EXPANDED_COL_WIDTHS = {
    COL_REF: 18, COL_NAME: 14, COL_CATEGORY: 45, COL_UNIT: 8, COL_TAGS: 40,
    COL_ATTR: 25, COL_VALUES: 50, COL_PLAN: 16, COL_REMARKS: 24, COL_SSKU: 32,
    COL_LAB: 25, COL_ORIGIN: 20, COL_UPLOAD_NAME: 30, COL_UPLOAD_ROW: 15,
}


def export_products_streaming(path, products, rows_for_product, total_products,
                              headers, sheet_name, progress_cb=None):
    """Stream a product export to disk, one product at a time.

    Uses openpyxl **write-only** mode so peak memory stays flat no matter
    how many rows are produced. ``rows_for_product(product)`` yields the
    rows for that product. ``progress_cb(done, total, rows_written)`` is
    invoked after each product for UI progress.
    Returns (products_done, rows_written).
    """
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(sheet_name)

    header_cells = []
    for h in headers:
        c = WriteOnlyCell(ws, value=h)
        c.fill = XL_HEADER_FILL
        c.font = XL_HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        header_cells.append(c)
    ws.append(header_cells)

    for idx, header in enumerate(headers, start=1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = EXPANDED_COL_WIDTHS.get(header, 14)
    ws.freeze_panes = "A2"

    rows_written = 0
    try:
        for i, product in enumerate(products, start=1):
            for row in rows_for_product(product):
                fill = XL_ALT_FILL if (rows_written % 2) else XL_WHITE_FILL
                cells = []
                for val in row:
                    c = WriteOnlyCell(ws, value=val)
                    c.fill = fill
                    c.font = XL_DATA_FONT
                    cells.append(c)
                ws.append(cells)
                rows_written += 1
            if progress_cb:
                progress_cb(i, total_products, rows_written)
        wb.save(path)
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return total_products, rows_written


# ==========================================================================
# Decoding engine
# ==========================================================================
def build_lookups():
    """Load every library fresh from disk into fast lookup structures."""
    L = {}

    # Stone name → 4-digit code (case-insensitive on the stone name)
    L["stone_by_name"] = {}
    for row in load_rows("stone"):
        name = norm(row.get("name"))
        if name and name not in L["stone_by_name"]:
            L["stone_by_name"][name] = {
                "code": clean(row.get("code")), "name": clean(row.get("name"))}

    # 2-digit shape codes
    L["shape_by_code"] = {}
    for row in load_rows("shape"):
        code = norm(row.get("code"))
        if code and code not in L["shape_by_code"]:
            L["shape_by_code"][code] = {
                "code": clean(row.get("code")), "name": clean(row.get("name")),
                "size_type": clean(row.get("size_type"))}

    # Full/variant shape names → 2-digit codes
    L["shape_by_fullname"] = {}
    for row in load_rows("fullname_shape"):
        full = norm(row.get("shape"))
        if full and full not in L["shape_by_fullname"]:
            L["shape_by_fullname"][full] = {
                "code": clean(row.get("code")), "name": clean(row.get("name")),
                "size_type": clean(row.get("size_type"))}

    # Polish type entries, indexed by normalised shape for fast filtering
    L["polish_by_shape"] = {}
    for row in load_rows("polish"):
        shape = norm(row.get("shape"))
        if not shape:
            continue
        entry = {
            "shape": shape,
            "ref_no": norm(row.get("ref_no")),
            "plan_no": norm(row.get("plan_no")),
            "remarks": norm(row.get("remarks")),
            "code": clean(row.get("code")),
            "name": clean(row.get("name")),
        }
        L["polish_by_shape"].setdefault(shape, []).append(entry)

    # Origin code → name
    L["origin_by_code"] = {}
    for row in load_rows("origin"):
        code = norm(row.get("code"))
        if code and code not in L["origin_by_code"]:
            L["origin_by_code"][code] = clean(row.get("name"))

    # Stone name → list of parent categories
    L["category_by_stone"] = {}
    for row in load_rows("category"):
        stone = norm(row.get("stone"))
        if stone:
            L["category_by_stone"].setdefault(stone, []).append(
                clean(row.get("parent")))

    return L


def match_polish(L, shape_raw, ref_no, plan_no, remarks):
    """Multi-level priority polish type match. Returns entry dict or None.

    Matches against the ORIGINAL upload shape value. Library fields that are
    blank act as wildcards; a blank upload field never matches a library
    entry that HAS a value for that field.
    """
    entries = L["polish_by_shape"].get(norm(shape_raw))
    if not entries:
        return None

    u = {"ref_no": norm(ref_no), "plan_no": norm(plan_no),
         "remarks": norm(remarks)}

    def field_ok(entry, field, contains=False):
        lib_val = entry[field]
        if not lib_val:               # blank library field = wildcard
            return True
        up_val = u[field]
        if not up_val:                # blank upload field ≠ valued library field
            return False
        if lib_val == up_val:
            return True
        if contains and field == "remarks":
            return lib_val in up_val or up_val in lib_val
        return False

    # Priority levels: required fields per level (shape already filtered).
    levels = [
        ("ref_no", "plan_no", "remarks"),   # 1
        ("ref_no", "plan_no"),              # 2
        ("ref_no", "remarks"),              # 3
        ("ref_no",),                        # 4
        ("plan_no", "remarks"),             # 5
        ("plan_no",),                       # 6
        ("remarks",),                       # 7
    ]
    for fields in levels:
        # Exact pass first; for levels involving Remarks a contains pass
        # follows when the exact pass found nothing.
        passes = [False, True] if "remarks" in fields else [False]
        for contains in passes:
            for entry in entries:
                # The upload row must actually have the fields this level
                # relies on — otherwise the level collapses to shape-only.
                if any(not u[f] for f in fields):
                    break
                if all(field_ok(entry, f, contains) for f in fields):
                    return entry

    # Level 8 — shape only: valid only if exactly ONE unique polish code
    # exists for this shape.
    codes = uniq(e["code"] for e in entries)
    if len(codes) == 1:
        return entries[0]
    return None


def parse_size(size_raw, height_raw):
    """Parse the Size / Height cells → (l, w, d, h) as strings.

    "X" format ("1.83X1.66") → L & W; a single value or a range → D.
    Mixed sizes never reach here — they are skipped during validation.
    All values are normalised to 2 decimal places via ``format_size_value``
    and feed the attribute columns directly.
    """
    s = clean(size_raw)
    l = w = d = ""
    if s:
        if re.search(r"[xX×]", s):
            parts = re.split(r"[xX×]", s)
            l = format_size_value(parts[0])
            w = format_size_value(parts[1]) if len(parts) > 1 else ""
        else:
            d = format_size_value(s)   # single value or range
    h = clean(height_raw)
    h = format_size_value(h) if h and is_number(h) else ""
    return l, w, d, h


def match_category(L, stone_name):
    """Find the Finish category path for a stone (all products are finished)."""
    parents = L["category_by_stone"].get(norm(stone_name), [])
    for parent in parents:
        if "finish" in parent.replace(" ", "").lower():
            return parent + " / " + stone_name
    return "All/ Stones/ Finish / " + stone_name


def dup_key(entry):
    """Duplicate-detection key for a decoded product entry."""
    return (norm(entry.get("internal_ref")), norm(entry.get("size")),
            norm(entry.get("origin_code")), norm(entry.get("stone")),
            norm(entry.get("polish_code")), norm(entry.get("shape_code")))


# --------------------------------------------------------------------------
# Project names / Internal Reference / duplicate resolution
# --------------------------------------------------------------------------
def is_project(remarks):
    """True when Remarks names one of the six shared-Ref.No projects."""
    return norm(remarks) in _PROJECT_LOOKUP


def build_internal_ref(ref_no, plan_no, remarks):
    """Internal Reference — ONLY from Ref.No (+ Plan No. for projects).

    Blank Ref.No → blank Internal Reference, whatever Order No., Remarks
    or Plan No. contain. For the six named projects the Plan No. is
    appended ("5711-104982-No 1") because those rows share one Ref.No.
    """
    ref = clean(ref_no)
    if not ref:
        return ""
    plan = clean(plan_no)
    if is_project(remarks) and plan:
        return f"{ref}-{plan}"
    return ref


def product_key(row):
    """Grouping key used to collapse duplicate upload rows.

    Project rows key on Ref.No + Plan No. + Variety + Shape + Size; other
    rows with a Ref.No drop the Plan No. Rows WITHOUT a Ref.No return None
    — they are never grouped, each one stays an individual product.
    """
    if not clean(row["ref_no"]):
        return None
    base = (norm(row["variety"]), norm(row["shape"]), norm(row["size"]))
    if is_project(row["remarks"]):
        return (norm(row["ref_no"]), norm(row["plan_no"])) + base
    return (norm(row["ref_no"]),) + base


def row_score(row):
    """How complete a row is — Ref.No + Remarks + Plan No. filled (0-3)."""
    return sum(1 for f in ("ref_no", "remarks", "plan_no") if clean(row[f]))


def resolve_duplicates(rows):
    """Collapse duplicate upload rows to the single best row of each group.

    The winner is the row with the most of Ref.No / Remarks / Plan No.
    filled (ties → the first row). Every LAB Certification in the group is
    collected onto the winner, as are all contributing upload row numbers.
    Losing rows are discarded silently — they are not skipped rows, they
    are the same product seen twice.
    Returns (kept_rows, duplicates_dropped).
    """
    groups = {}
    order = []
    kept = []
    for row in rows:
        key = product_key(row)
        if key is None:            # no Ref.No → always individual
            row["lab_all"] = clean(row["lab"])
            row["row_nums"] = str(row["row_num"])
            kept.append(row)
            continue
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(row)

    dropped = 0
    for key in order:
        members = groups[key]
        best = max(members, key=row_score)     # max() keeps the first on ties
        best["lab_all"] = ",".join(uniq(clean(m["lab"]) for m in members))
        best["row_nums"] = ",".join(uniq(str(m["row_num"]) for m in members))
        dropped += len(members) - 1
        kept.append(best)

    # Restore the original upload order so the preview reads naturally.
    kept.sort(key=lambda r: r["row_num"])
    return kept, dropped


def read_upload(path):
    """Read the 11-column invoice upload. Returns (rows, total_data_rows).

    Each row is a dict of cleaned strings plus its Excel row number.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    keys = ("item_no", "order_no", "ref_no", "lab", "origin", "variety",
            "shape", "size", "height", "plan_no", "remarks")
    rows = []
    header_seen = False
    for r_idx, raw in enumerate(ws.iter_rows(values_only=True), start=1):
        if raw is None:
            continue
        texts = [clean(c).lower() for c in raw]
        if not header_seen:
            # Treat the first row containing recognisable header words as
            # the header; anything before real data rows is also skipped.
            if any(t in ("variety", "shape", "ref.no", "order no.",
                         "item no.", "remarks") or "origin" in t
                   for t in texts if t):
                header_seen = True
                continue
            header_seen = True  # first row is data — fall through

        cells = list(raw)
        row = {}
        for i, k in enumerate(keys):
            row[k] = clean(cells[i]) if i < len(cells) else ""
        if not any(row.values()):
            continue
        row["row_num"] = r_idx
        rows.append(row)
    wb.close()
    return rows, len(rows)


def validate_row(row):
    """Pre-decode validation. Returns a skip reason, or "" when the row is OK.

    Steps 2-4: variety, shape and size must all be usable. Mixed sizes are
    rejected here — a mixed size cannot define a BOM.
    """
    variety = norm(row["variety"])
    if not variety or variety in INVALID_VARIETIES:
        return "Invalid variety"
    if not clean(row["shape"]):
        return "Invalid or missing shape"
    size = norm(row["size"])
    if size in MIXED_SIZES:
        return "Mixed size - cannot define BOM"
    if not size or size in INVALID_SIZES:
        return "Invalid or missing size"
    return ""


def process_rows(rows, upload_filename):
    """Run the full decode over every upload row.

    Order of work: clean/validate → collapse duplicate rows to the best row
    of each group → decode → enforce Internal Reference uniqueness → check
    the Previously Decoded Products library. Every surviving row becomes
    one INDIVIDUAL product; nothing is grouped or comma-separated.
    """
    L = build_lookups()
    existing_keys = {dup_key(e) for e in load_decoded()}

    products = []
    skipped = []

    def skip(row, reason):
        skipped.append({"filename": upload_filename, "row": row,
                        "reason": reason})

    # ---- Steps 1-4: clean (done on read) and validate -------------------
    valid_rows = []
    for row in rows:
        reason = validate_row(row)
        if reason:
            skip(row, reason)
        else:
            valid_rows.append(row)

    # ---- Step 12 (early): collapse duplicate upload rows ----------------
    kept_rows, duplicates_dropped = resolve_duplicates(valid_rows)

    # ---- Steps 5-11 + 13-14: decode each surviving row ------------------
    seen_refs = set()
    for row in kept_rows:
        # Step 5 — stone name
        stone = L["stone_by_name"].get(norm(row["variety"]))
        if not stone:
            skip(row, f"Stone name '{row['variety']}' not found in library")
            continue

        # Step 6 — shape (2-step: 2-digit codes, then full/variant names)
        shape = L["shape_by_code"].get(norm(row["shape"]))
        if not shape:
            shape = L["shape_by_fullname"].get(norm(row["shape"]))
        if not shape:
            skip(row, f"Shape '{row['shape']}' not found in library")
            continue

        # Step 7 — polish type (8-level priority match)
        polish = match_polish(L, row["shape"], row["ref_no"],
                              row["plan_no"], row["remarks"])
        if not polish:
            skip(row, f"Polish type not found for shape '{row['shape']}'")
            continue

        # Step 8 — origin
        origin_code = row["origin"]
        if origin_code:
            origin_name = L["origin_by_code"].get(norm(origin_code),
                                                  origin_code)
        else:
            origin_name = ""
        # Display origin for the export "Origin" column — resolved strictly
        # via the Origin Library (which holds both codes like "MM"/"CH/AU"
        # and full names like "Madagascar/Tanzania"). No match, or a blank
        # library name, → blank.
        origin_display = L["origin_by_code"].get(norm(origin_code), "") or "" \
            if origin_code else ""

        # Step 9 — 8-digit name code ("F" = finished product)
        polish_code = polish["code"] or "0"
        name_code = "F" + stone["code"] + polish_code + shape["code"]

        # Step 10 — size
        size_l, size_w, size_d, size_h = parse_size(row["size"], row["height"])

        # Step 11 — internal reference (ONLY from Ref.No)
        internal_ref = build_internal_ref(row["ref_no"], row["plan_no"],
                                          row["remarks"])

        # Step 14 — category, tags, SSKU with Size
        category = match_category(L, stone["name"])
        if polish_code == "0":
            tags = f"{stone['name']},{shape['name']}"
        else:
            tags = f"{stone['name']},{polish['name']},{shape['name']}"
        ssku = build_ssku_with_size(name_code, shape["size_type"], size_l,
                                    size_w, size_d, size_h, False)

        product = {
            "internal_ref": internal_ref,
            "name_code": name_code,
            "category": category,
            "tags": tags,
            "stone": stone["name"],
            "polish_code": polish_code,
            "polish_name": polish["name"],
            "shape_code": shape["code"],
            "shape_name": shape["name"],
            "size_type": shape["size_type"],
            "size": row["size"],
            "height": size_h,
            "ssku": ssku,
            "origin_code": origin_code,
            "origin_name": origin_name,
            "origin_display": origin_display,
            "plan_no": row["plan_no"],
            "remarks": row["remarks"],
            "order_no": row["order_no"],
            "lab_certification": row.get("lab_all", row["lab"]),
            "row_num": row["row_num"],
            "upload_filename": upload_filename,
            "upload_row_number": row.get("row_nums", str(row["row_num"])),
            "size_l": size_l,
            "size_w": size_w,
            "size_d": size_d,
            "size_h": size_h,
        }

        # Step 12 (final) — Internal Reference must be unique in the export.
        if internal_ref:
            if internal_ref in seen_refs:
                skip(row, "Duplicate Internal Reference")
                continue
            seen_refs.add(internal_ref)

        # Step 13 — already decoded in an earlier session? A blank Internal
        # Reference always counts as a new product.
        if internal_ref and dup_key(product) in existing_keys:
            skip(row, "Already decoded (duplicate in library)")
            continue

        products.append(product)

    return {
        "products": products,
        "skipped": skipped,
        "upload_filename": upload_filename,
        "duplicates_dropped": duplicates_dropped,
        "stats": {
            "matched": len(products),
            "skipped": len(skipped),
            "duplicates": duplicates_dropped,
        },
    }


# --------------------------------------------------------------------------
# Screen 3 — export row building (one product = one individual entry)
# --------------------------------------------------------------------------
def build_ssku_with_size(name_code, size_type, size_l, size_w, size_d,
                         size_h, size_mixed):
    """Build the 'SSKU with Size' string — sizes as company standard ranges.

    CASE 1  L&W, no height:   code-L-range,W-range
    CASE 2  L&W with height:  code-L-range,W-range,H-range
    CASE 3  D&H:              code-H-{range of Diameter} (actual Height IGNORED)
    CASE 4  D&H upload range: code-H-{range from floor of first number}
    CASE 5  Mixed:            code (no size suffix)
    All ranges come from ``size_to_range`` (floor-based X.00-X.99).
    """
    if size_mixed:
        return name_code
    l = size_to_range(size_l)
    w = size_to_range(size_w)
    h = size_to_range(size_h)
    d = size_to_range(size_d)
    if l or w:
        parts = []
        if l:
            parts.append(f"L-{l}")
        if w:
            parts.append(f"W-{w}")
        if h:
            parts.append(f"H-{h}")
        return name_code + "-" + ",".join(parts)
    if d:                       # Diameter maps onto the H axis; height ignored
        return name_code + "-H-" + d
    if h:                       # height-only edge case
        return name_code + "-H-" + h
    return name_code


def product_attributes(product):
    """Ordered attribute list for ONE product — all values single, never joined.

    Type → the size axes this product actually has → Treatment → Colour.
    """
    attrs = [("Type", "F")]
    if product["size_l"]:
        attrs.append(("Size-L", product["size_l"]))
    if product["size_w"]:
        attrs.append(("Size-W", product["size_w"]))
    if product["size_d"]:
        attrs.append(("Size-D", product["size_d"]))
    if product["size_h"]:
        attrs.append(("Size-H", product["size_h"]))
    attrs.append(("Treatment", "H"))
    attrs.append(("Colour", ""))
    return attrs


def _full_row(product, attr_name, attr_value):
    """The fully-populated 26-column row for one product."""
    return [
        product["internal_ref"],
        product["name_code"],
        product["category"],
        product["tags"],
        BOILER[COL_UNIT],
        attr_name, attr_value,
        BOILER["Sales"], BOILER["Purchase"], BOILER["Product Type"],
        BOILER["Invoicing Policy"], BOILER["Control Policy"],
        BOILER["Inventory tracking"], BOILER["Tracking"],
        BOILER["Sales Price"], BOILER["Sales Taxes"], BOILER["Cost"],
        BOILER["Purchase Taxes"], BOILER["Company"],
        product["ssku"],
        product["plan_no"],
        product["remarks"],
        product["lab_certification"],
        product["origin_display"],
        product["upload_filename"],
        product["upload_row_number"],
    ]


def _attr_only_row(attr_name, attr_value):
    """A continuation row: only columns 6-7 populated."""
    row = [""] * len(PRODUCT_HEADERS)
    row[5] = attr_name
    row[6] = attr_value
    return row


def product_rows_expanded(product):
    """Yield the expanded rows (one attribute per row) for one product."""
    attrs = product_attributes(product)
    first_name, first_value = attrs[0]
    yield _full_row(product, first_name, first_value)
    for name, value in attrs[1:]:
        yield _attr_only_row(name, value)


def product_rows_compact(product):
    """Yield the single compact row (all attributes joined) for one product."""
    attrs = product_attributes(product)
    yield _full_row(product,
                    ",".join(a for a, _ in attrs),
                    ",".join(v for _, v in attrs))


def build_skipped_rows(skipped):
    rows = []
    for item in skipped:
        r = item["row"]
        rows.append([
            r["row_num"], item["filename"], r["item_no"], r["order_no"],
            r["ref_no"], r["lab"], r["origin"], r["variety"], r["shape"],
            r["size"], r["height"], r["plan_no"], r["remarks"],
            item["reason"],
        ])
    return rows


# --------------------------------------------------------------------------
# Previously Decoded Products auto-save
# --------------------------------------------------------------------------
DECODED_KEYS = [k for _, k in DECODED_COLUMNS]


def autosave_decoded(products):
    """Append newly decoded products to the Previously Decoded library.

    Existing entries are never overwritten; duplicates (by key) are skipped.
    Returns the number of new entries added.
    """
    library = load_decoded()
    seen = {dup_key(e) for e in library}
    added = 0
    for p in products:
        entry = {k: p.get(k, "") for k in DECODED_KEYS}
        key = dup_key(entry)
        if key in seen:
            continue
        seen.add(key)
        library.append(entry)
        added += 1
    if added:
        save_decoded(library)
    return added


# ==========================================================================
# GUI widgets
# ==========================================================================
def _darken(hex_color, factor=0.85):
    h = hex_color.lstrip("#")
    r, g, b = (int(h[i:i + 2], 16) for i in (0, 2, 4))
    return "#%02x%02x%02x" % (int(r * factor), int(g * factor), int(b * factor))


class RoundButton(tk.Canvas):
    """A rounded-rectangle button drawn on a Canvas."""

    def __init__(self, parent, text, command, bg=HIGHLIGHT, fg="#FFFFFF",
                 width=170, height=42, radius=14, font=F_SUB, canvas_bg=BG):
        super().__init__(parent, width=width, height=height, bg=canvas_bg,
                         highlightthickness=0, bd=0)
        self.command = command
        self.bg, self.fg = bg, fg
        self.hover_bg = _darken(bg, 0.88)
        self.w, self.h, self.r = width, height, radius
        self._enabled = True
        self._rect = self._round_rect(1, 1, width - 1, height - 1, radius, fill=bg)
        self._text = self.create_text(width // 2, height // 2, text=text,
                                      fill=fg, font=font)
        self.bind("<Enter>", self._on_enter)
        self.bind("<Leave>", self._on_leave)
        self.bind("<Button-1>", self._on_click)
        self.configure(cursor="hand2")

    def _round_rect(self, x1, y1, x2, y2, r, **kw):
        pts = [x1 + r, y1, x2 - r, y1, x2, y1, x2, y1 + r, x2, y2 - r, x2, y2,
               x2 - r, y2, x1 + r, y2, x1, y2, x1, y2 - r, x1, y1 + r, x1, y1]
        return self.create_polygon(pts, smooth=True, **kw)

    def _on_enter(self, _):
        if self._enabled:
            self.itemconfig(self._rect, fill=self.hover_bg)

    def _on_leave(self, _):
        if self._enabled:
            self.itemconfig(self._rect, fill=self.bg)

    def _on_click(self, _):
        if self._enabled and self.command:
            self.command()

    def set_enabled(self, enabled):
        self._enabled = enabled
        self.itemconfig(self._rect, fill=self.bg if enabled else CARD)
        self.itemconfig(self._text, fill=self.fg if enabled else DIM)
        self.configure(cursor="hand2" if enabled else "arrow")


class ScrollFrame(tk.Frame):
    """A vertically scrollable frame (place widgets inside ``.inner``)."""

    def __init__(self, parent, bg=BG):
        super().__init__(parent, bg=bg)
        self.canvas = tk.Canvas(self, bg=bg, highlightthickness=0)
        self.vsb = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.inner = tk.Frame(self.canvas, bg=bg)
        self.inner.bind("<Configure>",
                        lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self._win = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")
        self.canvas.bind("<Configure>",
                         lambda e: self.canvas.itemconfig(self._win, width=e.width))
        self.canvas.configure(yscrollcommand=self.vsb.set)
        self.canvas.pack(side="left", fill="both", expand=True)
        self.vsb.pack(side="right", fill="y")
        self.canvas.bind("<Enter>", lambda e: self.canvas.bind_all("<MouseWheel>", self._wheel))
        self.canvas.bind("<Leave>", lambda e: self.canvas.unbind_all("<MouseWheel>"))

    def _wheel(self, event):
        self.canvas.yview_scroll(int(-event.delta / 120), "units")


def make_tree(parent, columns, headers, widths):
    """Create a horizontally + vertically scrollable Treeview."""
    frame = tk.Frame(parent, bg=BG)
    tree = ttk.Treeview(frame, columns=columns, show="headings")
    for c, h, w in zip(columns, headers, widths):
        tree.heading(c, text=h)
        tree.column(c, width=w, anchor="w", stretch=False)
    vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
    hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    tree.grid(row=0, column=0, sticky="nsew")
    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")
    frame.grid_rowconfigure(0, weight=1)
    frame.grid_columnconfigure(0, weight=1)
    return frame, tree


# ==========================================================================
# App
# ==========================================================================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("900x760")
        self.minsize(860, 620)
        self.configure(bg=BG)

        ensure_data_dir()
        self._set_icon()
        self._logo_img = self._load_logo()
        self._init_style()

        self.result = None

        self.container = tk.Frame(self, bg=BG)
        self.container.pack(fill="both", expand=True)
        self.container.grid_rowconfigure(0, weight=1)
        self.container.grid_columnconfigure(0, weight=1)

        self.frames = {}
        for Screen in (LibraryScreen, UploadScreen, ExportScreen):
            frame = Screen(self.container, self)
            self.frames[Screen.__name__] = frame
            frame.grid(row=0, column=0, sticky="nsew")
        self.show("LibraryScreen")

    def _set_icon(self):
        try:
            path = resource_path("icon.ico")
            if os.path.exists(path):
                self.iconbitmap(path)
        except Exception:
            pass

    def _load_logo(self):
        path = resource_path("logo.png")
        if not os.path.exists(path):
            return None
        try:
            if PIL_AVAILABLE:
                img = Image.open(path).resize((180, 64), Image.LANCZOS)
                return ImageTk.PhotoImage(img)
            return tk.PhotoImage(file=path)
        except Exception:
            return None

    def _init_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure("Treeview", background=BG, fieldbackground=BG,
                        foreground=TEXT, rowheight=24, font=F_BODY)
        style.configure("Treeview.Heading", background=CARD, foreground=TEXT,
                        font=F_SUB)
        style.map("Treeview", background=[("selected", HIGHLIGHT)],
                  foreground=[("selected", "#FFFFFF")])

    def show(self, name):
        frame = self.frames[name]
        if hasattr(frame, "on_show"):
            frame.on_show()
        frame.tkraise()


class BaseScreen(tk.Frame):
    """Shared header (centered logo + title) and footer for every screen."""

    def __init__(self, parent, app):
        super().__init__(parent, bg=BG)
        self.app = app
        self._build_header()
        self.body = tk.Frame(self, bg=BG)
        self.body.pack(fill="both", expand=True, padx=24, pady=8)
        self._build_footer()

    def _build_header(self):
        header = tk.Frame(self, bg=BG)
        header.pack(fill="x", pady=(16, 4))
        if self.app._logo_img is not None:
            tk.Label(header, image=self.app._logo_img, bg=BG).pack()
        else:
            tk.Label(header, text="ORAVA", font=(FONT, 22, "bold"),
                     fg=HIGHLIGHT, bg=BG).pack()
        tk.Label(header, text=APP_TITLE, font=F_TITLE, fg=TEXT, bg=BG).pack(pady=(6, 0))

    def _build_footer(self):
        footer = tk.Frame(self, bg=BG)
        footer.pack(fill="x", side="bottom")
        tk.Label(footer, text=FOOTER_TEXT, font=F_SMALL, fg=DIM, bg=BG).pack(pady=8)


# --------------------------------------------------------------------------
# Screen 1 — Library Manager
# --------------------------------------------------------------------------
class LibraryScreen(BaseScreen):
    def __init__(self, parent, app):
        super().__init__(parent, app)
        tk.Label(self.body, text="Reference Libraries", font=F_HEAD,
                 fg=TEXT, bg=BG).pack(anchor="w")
        tk.Label(self.body,
                 text="Manage the 8 libraries used to decode customised product specs.",
                 font=F_BODY, fg=DIM, bg=BG).pack(anchor="w", pady=(0, 10))

        scroller = ScrollFrame(self.body, bg=BG)
        scroller.pack(fill="both", expand=True)
        grid = scroller.inner
        grid.grid_columnconfigure(0, weight=1, uniform="lib")
        grid.grid_columnconfigure(1, weight=1, uniform="lib")

        # Per-library "entries loaded" labels, refreshed from disk on show and
        # after every Upload / Edit / Add / Delete All.
        self.count_lbls = {}
        for i, (key, cfg) in enumerate(LIBRARIES.items()):
            r, c = divmod(i, 2)
            self._make_card(grid, key, cfg).grid(row=r, column=c, sticky="nsew",
                                                 padx=6, pady=6)
        # Library 8 — special card
        r, c = divmod(len(LIBRARIES), 2)
        self._make_decoded_card(grid).grid(row=r, column=c, sticky="nsew",
                                           padx=6, pady=6)
        self.refresh_counts()

        nav = tk.Frame(self.body, bg=BG)
        nav.pack(fill="x", pady=(12, 0))
        RoundButton(nav, "Upload & Decode  →", self._go_next).pack(side="right")

    def _go_next(self):
        missing = [LIBRARIES[k]["title"] for k in ("stone", "shape", "polish")
                   if not load_rows(k)]
        if missing:
            messagebox.showwarning(
                "Libraries required",
                "Load these libraries before decoding:\n\n• "
                + "\n• ".join(missing))
            return
        self.app.show("UploadScreen")

    def on_show(self):
        # Counts are always read fresh from the JSON files.
        self.refresh_counts()

    def refresh_counts(self):
        """Reload every library from disk and update its entry-count label."""
        for key, lbl in self.count_lbls.items():
            count = len(load_decoded()) if key == "decoded" else len(load_rows(key))
            if count:
                lbl.configure(text=f"✔ {count} entries loaded", fg=SUCCESS_TEXT)
            else:
                lbl.configure(text="No data yet", fg=DIM)

    def _card_base(self, parent, title, desc):
        card = tk.Frame(parent, bg=CARD, padx=12, pady=10)
        tk.Label(card, text=title, font=F_SUB, fg=TEXT, bg=CARD,
                 anchor="w").pack(fill="x")
        tk.Label(card, text=desc, font=F_SMALL, fg=DIM, bg=CARD,
                 anchor="w", wraplength=360, justify="left").pack(fill="x")
        count_lbl = tk.Label(card, text="", font=F_SMALL, fg=DIM, bg=CARD,
                             anchor="w")
        count_lbl.pack(fill="x", pady=(2, 0))
        btns = tk.Frame(card, bg=CARD)
        btns.pack(fill="x", pady=(8, 0))
        return card, count_lbl, btns

    @staticmethod
    def _mk_btn(parent, text, cmd, color, tcolor="#FFFFFF"):
        return tk.Button(parent, text=text, command=cmd, font=F_SMALL,
                         bg=color, fg=tcolor, relief="flat", bd=0,
                         cursor="hand2", padx=8, pady=4,
                         activebackground=color, activeforeground=tcolor)

    def _make_card(self, parent, key, cfg):
        card, count_lbl, btns = self._card_base(parent, cfg["title"], cfg["desc"])
        self.count_lbls[key] = count_lbl
        self._mk_btn(btns, "Upload", lambda: self._upload(key),
                     HIGHLIGHT).pack(side="left", padx=2)
        self._mk_btn(btns, "Edit",
                     lambda: EditDialog(self.app, key, on_change=self.refresh_counts),
                     CARD, TEXT).pack(side="left", padx=2)
        self._mk_btn(btns, "Download", lambda: self._download(key),
                     CARD, TEXT).pack(side="left", padx=2)
        self._mk_btn(btns, "Delete All", lambda: self._delete_all(key),
                     DANGER).pack(side="right", padx=2)
        return card

    def _make_decoded_card(self, parent):
        card, count_lbl, btns = self._card_base(
            parent, "Previously Decoded Products",
            "Stores all decoded products to prevent duplicates. "
            "Auto-populated on export.")
        self.count_lbls["decoded"] = count_lbl
        self._mk_btn(btns, "Export", self._export_decoded,
                     INFO).pack(side="left", padx=2)
        self._mk_btn(btns, "Search", self._search_decoded,
                     SUCCESS).pack(side="left", padx=2)
        self._mk_btn(btns, "Delete All", self._delete_decoded,
                     DANGER).pack(side="right", padx=2)
        return card

    # ---- standard library actions -------------------------------------
    def _upload(self, key):
        path = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")])
        if not path:
            return
        try:
            _, added, _ = import_excel(key, path)
        except Exception as exc:
            messagebox.showerror("Upload failed", str(exc))
            return
        self.refresh_counts()
        messagebox.showinfo("Upload complete",
                            f"{LIBRARIES[key]['title']}\n\n"
                            f"Added {added} rows (duplicates kept).")

    def _download(self, key):
        path = filedialog.asksaveasfilename(
            title="Save library", defaultextension=".xlsx",
            initialfile=f"{LIBRARIES[key]['title']}.xlsx",
            filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            export_library_excel(key, path)
        except Exception as exc:
            messagebox.showerror("Download failed", str(exc))
            return
        messagebox.showinfo("Saved", f"Exported to:\n{path}")

    def _delete_all(self, key):
        if messagebox.askyesno("Delete All",
                               f"Delete ALL entries in {LIBRARIES[key]['title']}?\n"
                               "This cannot be undone."):
            save_rows(key, [])
            self.refresh_counts()
            messagebox.showinfo("Deleted", "All entries removed.")

    # ---- Previously Decoded Products actions ---------------------------
    def _export_decoded(self):
        if not load_decoded():
            messagebox.showinfo("Export", "The library is empty.")
            return
        path = filedialog.asksaveasfilename(
            title="Export Previously Decoded Products",
            defaultextension=".xlsx",
            initialfile="Previously Decoded Products.xlsx",
            filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            export_decoded_excel(path)
        except Exception as exc:
            messagebox.showerror("Export failed", str(exc))
            return
        messagebox.showinfo("Saved", f"Exported to:\n{path}")

    def _search_decoded(self):
        DecodedSearchDialog(self.app)

    def _delete_decoded(self):
        if messagebox.askyesno(
                "Delete All",
                "Delete ALL entries in Previously Decoded Products?\n"
                "Duplicate detection will restart from empty.\n"
                "This cannot be undone."):
            save_decoded([])
            self.refresh_counts()
            messagebox.showinfo("Deleted", "All entries removed.")


class DecodedSearchDialog(tk.Toplevel):
    """Universal search across ALL columns of Previously Decoded Products."""

    def __init__(self, app):
        super().__init__(app)
        self.title("Search — Previously Decoded Products")
        self.configure(bg=BG)
        self.geometry("1000x560")
        self.minsize(700, 400)
        self.transient(app)
        self.grab_set()

        tk.Label(self, text="Previously Decoded Products — Universal Search",
                 font=F_HEAD, fg=TEXT, bg=BG).pack(pady=(12, 6))

        top = tk.Frame(self, bg=BG)
        top.pack(fill="x", padx=16)
        tk.Label(top, text="Search any value:", font=F_BODY, fg=DIM,
                 bg=BG).pack(side="left")
        self.search_var = tk.StringVar()
        entry = tk.Entry(top, textvariable=self.search_var, font=F_BODY,
                         bg=CARD, fg=TEXT, relief="flat")
        entry.pack(side="left", fill="x", expand=True, padx=8, ipady=3)
        entry.bind("<Return>", lambda e: self.refresh())
        tk.Button(top, text="Search", command=self.refresh, font=F_SUB,
                  bg=HIGHLIGHT, fg="#FFFFFF", relief="flat", cursor="hand2",
                  padx=14, pady=4).pack(side="left")

        self.count_lbl = tk.Label(self, text="", font=F_SMALL, fg=DIM, bg=BG)
        self.count_lbl.pack(anchor="w", padx=16, pady=(6, 0))

        cols = [k for _, k in DECODED_COLUMNS]
        heads = [d for d, _ in DECODED_COLUMNS]
        table, self.tree = make_tree(self, cols, heads, [130] * len(cols))
        table.pack(fill="both", expand=True, padx=16, pady=8)

        self.rows = load_decoded()
        self.refresh()
        entry.focus_set()

    def refresh(self):
        term = self.search_var.get().replace("\xa0", " ").strip().lower()
        self.tree.delete(*self.tree.get_children())
        keys = [k for _, k in DECODED_COLUMNS]
        shown = 0
        for row in self.rows:
            values = [row.get(k, "") for k in keys]
            if term and not any(term in str(v).lower() for v in values):
                continue
            self.tree.insert("", "end", values=values)
            shown += 1
            if shown >= PREVIEW_LIMIT:
                break
        total = len(self.rows)
        note = f"{shown} matching entries (of {total} stored)"
        if shown >= PREVIEW_LIMIT:
            note += f" — showing first {PREVIEW_LIMIT}"
        self.count_lbl.configure(text=note)


# --------------------------------------------------------------------------
# Edit dialog (search / add / edit per row / delete per row)
# --------------------------------------------------------------------------
class EditDialog(tk.Toplevel):
    def __init__(self, app, lib_key, on_change=None):
        super().__init__(app)
        self.lib_key = lib_key
        # Called after any add / edit / delete so the caller (Screen 1) can
        # refresh its entry-count label.
        self.on_change = on_change
        self.cfg = LIBRARIES[lib_key]
        self.title("Edit — " + self.cfg["title"])
        self.configure(bg=BG)
        self.geometry("760x520")
        self.minsize(560, 380)
        self.transient(app)
        self.grab_set()

        tk.Label(self, text=self.cfg["title"], font=F_HEAD, fg=TEXT,
                 bg=BG).pack(pady=(12, 6))

        top = tk.Frame(self, bg=BG)
        top.pack(fill="x", padx=16)
        tk.Label(top, text="Search:", font=F_BODY, fg=DIM, bg=BG).pack(side="left")
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *_: self.refresh())
        tk.Entry(top, textvariable=self.search_var, font=F_BODY, bg=CARD, fg=TEXT,
                 relief="flat").pack(side="left", fill="x", expand=True, padx=8, ipady=3)

        cols = [k for _, k in self.cfg["columns"]]
        heads = [d for d, _ in self.cfg["columns"]]
        table, self.tree = make_tree(self, cols, heads, [150] * len(cols))
        table.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<Double-1>", lambda e: self.edit_row())

        actions = tk.Frame(self, bg=BG)
        actions.pack(fill="x", padx=16, pady=(0, 12))
        for text, cmd, color, tc in (
                ("Add", self.add_row, HIGHLIGHT, "#FFFFFF"),
                ("Edit", self.edit_row, CARD, TEXT),
                ("Delete", self.delete_row, CARD, TEXT)):
            tk.Button(actions, text=text, command=cmd, font=F_BODY, bg=color, fg=tc,
                      relief="flat", cursor="hand2", padx=12, pady=4).pack(side="left", padx=3)

        self.rows = load_rows(lib_key)
        self.refresh()

    def refresh(self):
        term = self.search_var.get().strip().lower()
        self.tree.delete(*self.tree.get_children())
        keys = [k for _, k in self.cfg["columns"]]
        for idx, row in enumerate(self.rows):
            values = [row.get(k, "") for k in keys]
            if term and not any(term in str(v).lower() for v in values):
                continue
            self.tree.insert("", "end", iid=str(idx), values=values)

    def _selected_index(self):
        sel = self.tree.selection()
        return int(sel[0]) if sel else None

    def _persist(self):
        save_rows(self.lib_key, self.rows)
        self.refresh()
        if self.on_change:
            self.on_change()

    def add_row(self):
        RowEditor(self, self.cfg["columns"], None, self._on_add)

    def _on_add(self, values):
        keys = [k for _, k in self.cfg["columns"]]
        # Duplicates are allowed — libraries keep spelling variations.
        self.rows.append({k: values[k] for k in keys})
        self._persist()

    def edit_row(self):
        idx = self._selected_index()
        if idx is None:
            messagebox.showinfo("Edit", "Select a row first.", parent=self)
            return
        RowEditor(self, self.cfg["columns"], self.rows[idx],
                  lambda vals: self._on_edit(idx, vals))

    def _on_edit(self, idx, values):
        keys = [k for _, k in self.cfg["columns"]]
        self.rows[idx] = {k: values[k] for k in keys}
        self._persist()

    def delete_row(self):
        idx = self._selected_index()
        if idx is None:
            messagebox.showinfo("Delete", "Select a row first.", parent=self)
            return
        if messagebox.askyesno("Delete row", "Delete the selected row?", parent=self):
            self.rows.pop(idx)
            self._persist()


class RowEditor(tk.Toplevel):
    def __init__(self, parent, columns, existing, on_save):
        super().__init__(parent)
        self.on_save = on_save
        self.title("Edit Row" if existing else "Add Row")
        self.configure(bg=BG)
        self.transient(parent)
        self.grab_set()
        self.resizable(False, False)

        self.vars = {}
        form = tk.Frame(self, bg=BG)
        form.pack(padx=20, pady=16)
        for i, (disp, k) in enumerate(columns):
            tk.Label(form, text=disp, font=F_BODY, fg=TEXT, bg=BG).grid(
                row=i, column=0, sticky="w", pady=4, padx=(0, 8))
            var = tk.StringVar(value=(existing or {}).get(k, ""))
            tk.Entry(form, textvariable=var, font=F_BODY, width=34, bg=CARD, fg=TEXT,
                     relief="flat").grid(row=i, column=1, ipady=3)
            self.vars[k] = var

        btns = tk.Frame(self, bg=BG)
        btns.pack(pady=(0, 14))
        tk.Button(btns, text="Save", command=self._save, font=F_SUB, bg=HIGHLIGHT,
                  fg="#FFFFFF", relief="flat", cursor="hand2", padx=14, pady=5
                  ).pack(side="left", padx=6)
        tk.Button(btns, text="Cancel", command=self.destroy, font=F_SUB, bg=CARD,
                  fg=TEXT, relief="flat", cursor="hand2", padx=14, pady=5
                  ).pack(side="left", padx=6)

    def _save(self):
        values = {k: clean(v.get()) for k, v in self.vars.items()}
        if not any(values.values()):
            messagebox.showwarning("Empty", "Enter at least one value.", parent=self)
            return
        self.on_save(values)
        self.destroy()


# --------------------------------------------------------------------------
# Screen 2 — Upload & Process
# --------------------------------------------------------------------------
class UploadScreen(BaseScreen):
    def __init__(self, parent, app):
        super().__init__(parent, app)
        tk.Label(self.body, text="Upload & Process", font=F_HEAD,
                 fg=TEXT, bg=BG).pack(anchor="w")
        tk.Label(self.body,
                 text="Excel with columns: Item No., Order No., Ref.No, LAB "
                      "Certification, Origin, Variety, Shape, Size, Height, "
                      "Plan No., Remarks.",
                 font=F_BODY, fg=DIM, bg=BG).pack(anchor="w", pady=(0, 10))

        pick = tk.Frame(self.body, bg=CARD, padx=14, pady=12)
        pick.pack(fill="x")
        self.path_var = tk.StringVar(value="No file selected.")
        tk.Label(pick, textvariable=self.path_var, font=F_BODY, fg=DIM, bg=CARD,
                 anchor="w").pack(fill="x")
        prow = tk.Frame(pick, bg=CARD)
        prow.pack(fill="x", pady=(10, 0))
        tk.Button(prow, text="Choose File…", command=self.choose_file, font=F_SUB,
                  bg="#FFFFFF", fg=TEXT, relief="solid", bd=1, cursor="hand2",
                  padx=12, pady=6).pack(side="left")
        RoundButton(prow, "Process", self.process, canvas_bg=CARD,
                    width=130).pack(side="left", padx=10)

        self.summary_lbl = tk.Label(self.body, text="", font=F_BODY, fg=TEXT,
                                    bg=BG, justify="left")
        self.summary_lbl.pack(anchor="w", pady=8)

        cols = ("row", "refno", "variety", "shape", "size", "code", "status",
                "reason")
        heads = ("Row", "Internal Ref", "Variety", "Shape", "Size", "8-digit",
                 "Status", "Reason")
        table, self.tree = make_tree(self.body, cols, heads,
                                     (50, 110, 150, 110, 110, 90, 90, 260))
        table.pack(fill="both", expand=True)
        self.tree.tag_configure("ok", foreground=SUCCESS_TEXT)
        self.tree.tag_configure("bad", foreground=DANGER)

        nav = tk.Frame(self.body, bg=BG)
        nav.pack(fill="x", side="bottom", pady=(10, 0))
        tk.Button(nav, text="←  Libraries",
                  command=lambda: self.app.show("LibraryScreen"),
                  font=F_BODY, bg=CARD, fg=TEXT, relief="flat", cursor="hand2",
                  padx=10, pady=6).pack(side="left")
        self.next_btn = RoundButton(nav, "Export  →",
                                    lambda: self.app.show("ExportScreen"),
                                    width=150)
        self.next_btn.pack(side="right")
        self.next_btn.set_enabled(False)

        self.selected_path = None

    def on_show(self):
        pass

    def choose_file(self):
        path = filedialog.askopenfilename(
            title="Select customer invoice Excel file",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")])
        if path:
            self.selected_path = path
            self.path_var.set(path)

    def process(self):
        if not self.selected_path:
            messagebox.showinfo("Upload", "Choose a file first.")
            return
        try:
            rows, total = read_upload(self.selected_path)
        except Exception as exc:
            messagebox.showerror("Read failed", str(exc))
            return
        if not rows:
            messagebox.showwarning("Empty", "No data rows found.")
            return

        filename = os.path.basename(self.selected_path)
        self.app.result = process_rows(rows, filename)
        s = self.app.result["stats"]
        self.summary_lbl.configure(
            text=(f"Loaded {total} rows from {filename}.\n"
                  f"Decoded: {s['matched']} individual products  •  "
                  f"{s['duplicates']} duplicate rows merged.\n"
                  f"Mismatched / skipped: {s['skipped']}"))

        # Preview: matched first, then mismatched (capped for performance).
        self.tree.delete(*self.tree.get_children())
        shown = 0
        for p in self.app.result["products"]:
            if shown >= PREVIEW_LIMIT:
                break
            self.tree.insert("", "end", values=(
                p["row_num"], p["internal_ref"], p["stone"], p["shape_code"],
                p["size"], p["name_code"], "matched", ""), tags=("ok",))
            shown += 1
        for item in self.app.result["skipped"]:
            if shown >= PREVIEW_LIMIT:
                break
            r = item["row"]
            self.tree.insert("", "end", values=(
                r["row_num"], r["ref_no"], r["variety"], r["shape"],
                r["size"], "", "mismatched", item["reason"]), tags=("bad",))
            shown += 1

        self.next_btn.set_enabled(True)
        self.app.frames["ExportScreen"].load()


# --------------------------------------------------------------------------
# Export progress popup
# --------------------------------------------------------------------------
class ProgressDialog(tk.Toplevel):
    """Modal progress popup with a bar + live product / row counters."""

    def __init__(self, app, title="Exporting…", maximum=100):
        super().__init__(app)
        self.app = app
        self.title(title)
        self.configure(bg=BG)
        self.resizable(False, False)
        self.transient(app)
        self.geometry("380x150")

        tk.Label(self, text=title, font=F_SUB, fg=TEXT, bg=BG).pack(pady=(18, 8))
        self.status = tk.Label(self, text="Preparing…", font=F_BODY, fg=TEXT, bg=BG)
        self.status.pack()
        self.rows_lbl = tk.Label(self, text="Total rows written: 0",
                                 font=F_SMALL, fg=DIM, bg=BG)
        self.rows_lbl.pack(pady=(2, 10))

        self.bar = ttk.Progressbar(self, orient="horizontal", mode="determinate",
                                   length=320, maximum=max(maximum, 1))
        self.bar.pack(pady=(0, 14))

        # Keep the popup on top; ignore the close button during export.
        self.protocol("WM_DELETE_WINDOW", lambda: None)
        self.grab_set()
        self.update()

    def update_progress(self, done, total, rows_written):
        self.bar["maximum"] = max(total, 1)
        self.bar["value"] = done
        self.status.configure(text=f"Exporting product {done} of {total}…")
        self.rows_lbl.configure(text=f"Total rows written: {rows_written:,}")
        # Full event pump keeps the GUI responsive during a long export.
        self.app.update()

    def done(self):
        try:
            self.grab_release()
        except Exception:
            pass
        self.destroy()


# --------------------------------------------------------------------------
# Screen 3 — Export (26 columns, one individual product per entry)
# --------------------------------------------------------------------------
class ExportScreen(BaseScreen):
    """Final screen: preview every decoded product and export it.

    Each decoded upload row is one INDIVIDUAL product — there is no
    grouping and no comma-separated attribute values anywhere here.
    """

    def __init__(self, parent, app):
        super().__init__(parent, app)
        top = tk.Frame(self.body, bg=BG)
        top.pack(fill="x")
        tk.Label(top, text="Export  (26 columns)", font=F_HEAD,
                 fg=TEXT, bg=BG).pack(side="left")
        # Packed right-to-left, so they read: Expanded | Compact | Skipped.
        RoundButton(top, "Export Skipped", self.export_skipped, bg=WARNING,
                    width=140).pack(side="right", padx=(6, 0))
        RoundButton(top, "Export Compact", self.export_compact, bg=INFO,
                    width=140).pack(side="right", padx=(6, 0))
        RoundButton(top, "Export Expanded", self.export_expanded, bg=SUCCESS,
                    width=140).pack(side="right")

        card = tk.Frame(self.body, bg=CARD, padx=14, pady=10)
        card.pack(fill="x", pady=(8, 0))
        self.summary_lbl = tk.Label(card, text="Process a file first.",
                                    font=F_SUB, fg=TEXT, bg=CARD, anchor="w")
        self.summary_lbl.pack(fill="x")
        self.note_lbl = tk.Label(card, text="", font=F_SMALL, fg=DIM, bg=CARD,
                                 anchor="w")
        self.note_lbl.pack(fill="x")

        widths = ([120, 90, 240, 200, 55, 110, 110] + [80] * 11
                  + [260, 120, 140, 130, 110, 180, 110])
        table, self.tree = make_tree(self.body, PRODUCT_HEADERS,
                                     PRODUCT_HEADERS, widths)
        table.pack(fill="both", expand=True, pady=8)

        nav = tk.Frame(self.body, bg=BG)
        nav.pack(fill="x", side="bottom", pady=(8, 0))
        tk.Button(nav, text="←  Upload",
                  command=lambda: self.app.show("UploadScreen"),
                  font=F_BODY, bg=CARD, fg=TEXT, relief="flat", cursor="hand2",
                  padx=10, pady=6).pack(side="left")
        tk.Button(nav, text="Close", command=self.app.destroy, font=F_BODY,
                  bg=DANGER, fg="#FFFFFF", relief="flat", cursor="hand2",
                  padx=14, pady=6).pack(side="right")

    def on_show(self):
        self.load()

    def load(self):
        self.tree.delete(*self.tree.get_children())
        self.note_lbl.configure(text="")
        if not self.app.result:
            self.summary_lbl.configure(text="Process a file first.")
            return
        s = self.app.result["stats"]
        self.summary_lbl.configure(
            text=f"Total Products: {s['matched']}   |   Skipped: {s['skipped']}")

        # Only the capped preview slice is rendered so the Treeview never
        # freezes on very large result sets — the export always writes all rows.
        shown = 0
        capped = False
        for product in self.app.result["products"]:
            if capped:
                break
            for row in product_rows_expanded(product):
                if shown >= PREVIEW_LIMIT:
                    capped = True
                    break
                self.tree.insert("", "end", values=row)
                shown += 1
        if capped:
            self.note_lbl.configure(
                text=(f"Preview shows the first {PREVIEW_LIMIT:,} rows only — "
                      f"the export writes everything."))
        else:
            self.note_lbl.configure(
                text=f"{s['duplicates']} duplicate upload rows merged  •  "
                     f"one row per attribute.")

    def _autosave(self):
        added = autosave_decoded(self.app.result["products"])
        self.app.frames["LibraryScreen"].refresh_counts()
        return added

    def export_expanded(self):
        """Expanded format — one row per attribute, per product."""
        self._run_export(
            rows_for_product=product_rows_expanded,
            title="Export Expanded",
            initialfile="Customised Product Spec - Decoded.xlsx",
            sheet_name="Decoded",
            progress_title="Exporting Decoded Products…")

    def export_compact(self):
        """Compact format — all attributes on ONE row per product."""
        self._run_export(
            rows_for_product=product_rows_compact,
            title="Export Compact",
            initialfile="Customised Product Spec - Compact.xlsx",
            sheet_name="Compact",
            progress_title="Exporting Compact Products…")

    def _run_export(self, rows_for_product, title, initialfile, sheet_name,
                    progress_title):
        """Shared streamed-export driver (file picker + progress popup + errors)."""
        if not self.app.result or not self.app.result["products"]:
            messagebox.showinfo("Export", "Nothing to export yet.")
            return
        products = self.app.result["products"]
        path = filedialog.asksaveasfilename(
            title=title, defaultextension=".xlsx",
            initialfile=initialfile,
            filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return

        dialog = ProgressDialog(self.app, title=progress_title,
                                maximum=len(products))
        try:
            done, rows_written = export_products_streaming(
                path, products, rows_for_product, len(products),
                PRODUCT_HEADERS, sheet_name,
                progress_cb=dialog.update_progress)
        except MemoryError:
            dialog.done()
            messagebox.showerror(
                "Out of memory",
                "Ran out of memory while exporting.\n\n"
                "Try processing fewer rows and export in batches.")
            return
        except Exception as exc:
            dialog.done()
            messagebox.showerror("Export failed",
                                 f"The export could not be completed:\n\n{exc}")
            return
        dialog.done()
        added = self._autosave()
        messagebox.showinfo(
            "Exported",
            f"Saved to:\n{path}\n\n"
            f"Products exported: {done:,}\n"
            f"Total rows written: {rows_written:,}\n"
            f"{added} products added to Previously Decoded Products library.")

    def export_skipped(self):
        if not self.app.result:
            messagebox.showinfo("Export", "Process a file first.")
            return
        skipped = self.app.result["skipped"]
        if not skipped:
            messagebox.showinfo("Export Skipped", "No skipped rows 🎉")
            return
        path = filedialog.asksaveasfilename(
            title="Export Skipped", defaultextension=".xlsx",
            initialfile="Customised Product Spec - Skipped.xlsx",
            filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            write_styled_workbook(path, SKIPPED_HEADERS,
                                  build_skipped_rows(skipped), "Skipped")
        except Exception as exc:
            messagebox.showerror("Export failed", str(exc))
            return
        messagebox.showinfo("Exported", f"Saved to:\n{path}")


# --------------------------------------------------------------------------
if __name__ == "__main__":
    App().mainloop()
